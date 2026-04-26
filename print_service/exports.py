import io
import os
from datetime import datetime


def _ensure_openpyxl():
  import openpyxl
  from openpyxl.drawing.image import Image as XLImage
  from openpyxl.styles import Alignment, Font, PatternFill
  return openpyxl, XLImage, Alignment, Font, PatternFill


def export_excel(rows, totals, company_name="Librería Virgen de la Puerta", logo_path=None, title="Reporte de Impresiones"):
  openpyxl, XLImage, Alignment, Font, PatternFill = _ensure_openpyxl()

  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Impresiones"

  ws.merge_cells("A1:L1")
  ws["A1"] = company_name
  ws["A1"].font = Font(bold=True, size=14, color="0F172A")
  ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

  ws.merge_cells("A2:L2")
  ws["A2"] = title
  ws["A2"].font = Font(bold=True, size=12, color="111827")
  ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

  now = datetime.now().astimezone().isoformat(timespec="seconds")
  ws.merge_cells("A3:L3")
  ws["A3"] = now
  ws["A3"].font = Font(size=10, color="6B7280")
  ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

  if logo_path and os.path.exists(logo_path):
    try:
      img = XLImage(logo_path)
      img.width = 70
      img.height = 70
      ws.add_image(img, "L1")
    except Exception:
      pass

  ws.append([])

  headers = [
    "ID",
    "Fecha creación",
    "Fecha completado",
    "Impresora",
    "Usuario",
    "Documento",
    "Tipo",
    "Confirmadas",
    "Estimadas (copias)",
    "Copias",
    "Estado",
    "Error",
  ]

  hfill = PatternFill("solid", fgColor="0F172A")
  hfont = Font(bold=True, color="FFFFFF")
  header_row = ws.max_row + 1
  for col, h in enumerate(headers, 1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.fill = hfill
    c.font = hfont
    c.alignment = Alignment(horizontal="center")

  data_start = header_row + 1
  for i, r in enumerate(rows, data_start):
    ws.cell(i, 1, r.get("id"))
    ws.cell(i, 2, r.get("ts_created"))
    ws.cell(i, 3, r.get("ts_completed"))
    ws.cell(i, 4, r.get("printer_name"))
    ws.cell(i, 5, r.get("user_id"))
    ws.cell(i, 6, r.get("document"))
    ws.cell(i, 7, r.get("print_type"))
    ws.cell(i, 8, r.get("pages"))
    ws.cell(i, 9, r.get("pages_estimated"))
    ws.cell(i, 10, r.get("copies_requested"))
    ws.cell(i, 11, r.get("status"))
    ws.cell(i, 12, r.get("error_code"))

  widths = [8, 22, 22, 28, 20, 42, 10, 12, 14, 10, 12, 18]
  for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

  last_row = len(rows) + header_row
  ws.append([])
  ws.append(["", "", "", "", "", "TOTAL CONFIRMADAS", "", f"=SUM(H{data_start}:H{last_row})"])
  ws.append(["", "", "", "", "", "TOTAL ESTIMADAS", "", f"=SUM(I{data_start}:I{last_row})"])
  ws.append(["", "", "", "", "", "TOTAL BN (CONF)", "", totals.get("pages_bn", 0)])
  ws.append(["", "", "", "", "", "TOTAL COLOR (CONF)", "", totals.get("pages_color", 0)])
  ws.append(["", "", "", "", "", "TRABAJOS OK", "", totals.get("completed_jobs", 0)])
  ws.append(["", "", "", "", "", "TRABAJOS FALLIDOS", "", totals.get("failed_jobs", 0)])

  meta = wb.create_sheet("Columnas")
  meta["A1"] = "Columna"
  meta["B1"] = "Descripción"
  meta["A1"].font = Font(bold=True)
  meta["B1"].font = Font(bold=True)
  cols = [
    ("ID", "ID del trabajo de impresión."),
    ("Fecha creación", "Timestamp de creación del job."),
    ("Fecha completado", "Timestamp de completado del job."),
    ("Impresora", "Nombre de impresora detectada."),
    ("Usuario", "Usuario/host asociado."),
    ("Documento", "Nombre del documento (si disponible)."),
    ("Tipo", "BN o Color (según detección)."),
    ("Confirmadas", "Páginas confirmadas por job."),
    ("Estimadas (copias)", "Estimación de páginas/copias (si aplica)."),
    ("Copias", "Copias solicitadas."),
    ("Estado", "Estado del job."),
    ("Error", "Código/error del job.")
  ]
  for i, (c, d) in enumerate(cols, 2):
    meta.cell(i, 1, c)
    meta.cell(i, 2, d)

  bio = io.BytesIO()
  wb.save(bio)
  bio.seek(0)
  return bio.read()


def _ensure_reportlab():
  from reportlab.lib import colors
  from reportlab.lib.pagesizes import A4
  from reportlab.lib.styles import getSampleStyleSheet
  from reportlab.lib.units import cm
  from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
  return colors, A4, getSampleStyleSheet, cm, Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def export_pdf(rows, totals, company_name="Librería Virgen de la Puerta", logo_path=None, title="Reporte de Impresiones"):
  colors, A4, getSampleStyleSheet, cm, Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle = _ensure_reportlab()

  buf = io.BytesIO()
  doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.4 * cm, rightMargin=1.4 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm)
  styles = getSampleStyleSheet()

  story = []
  top = []
  if logo_path and os.path.exists(logo_path):
    try:
      top.append(Image(logo_path, width=2.2 * cm, height=2.2 * cm))
    except Exception:
      top.append(Paragraph("", styles["Normal"]))
  else:
    top.append(Paragraph("", styles["Normal"]))
  now = datetime.now().astimezone().isoformat(timespec="seconds")
  top.append(Paragraph(f"<b>{company_name}</b><br/>{title}<br/><font size=9>{now}</font>", styles["Normal"]))
  ttop = Table([top], colWidths=[2.6 * cm, 14.2 * cm])
  ttop.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
  story.append(ttop)
  story.append(Spacer(1, 0.4 * cm))

  summary_data = [
    ["Confirmadas", totals.get("pages_total", 0), "Estimadas", totals.get("pages_total_estimated", 0), "OK", totals.get("completed_jobs", 0)],
    ["BN (conf)", totals.get("pages_bn", 0), "Color (conf)", totals.get("pages_color", 0), "Fallidos", totals.get("failed_jobs", 0)],
  ]
  tsum = Table(summary_data, colWidths=[3.0 * cm, 2.5 * cm, 2.0 * cm, 2.5 * cm, 2.0 * cm, 2.5 * cm])
  tsum.setStyle(
    TableStyle(
      [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
      ]
    )
  )
  story.append(tsum)
  story.append(Spacer(1, 0.5 * cm))

  headers = ["ID", "Creado", "Completado", "Impresora", "Usuario", "Tipo", "Conf", "Est", "Copias"]
  data = [headers]
  for r in rows[:500]:
    data.append(
      [
        str(r.get("id", "")),
        (r.get("ts_created") or "")[:19],
        (r.get("ts_completed") or "")[:19],
        (r.get("printer_name") or "")[:22],
        (r.get("user_id") or "")[:18],
        r.get("print_type") or "",
        str(r.get("pages") or 0),
        str(r.get("pages_estimated") or 0),
        str(r.get("copies_requested") or 1),
      ]
    )
  table = Table(data, repeatRows=1, colWidths=[1.2 * cm, 2.6 * cm, 2.6 * cm, 3.0 * cm, 3.0 * cm, 1.4 * cm, 1.4 * cm, 1.4 * cm, 1.4 * cm])
  table.setStyle(
    TableStyle(
      [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
      ]
    )
  )
  story.append(table)

  doc.build(story)
  buf.seek(0)
  return buf.read()

